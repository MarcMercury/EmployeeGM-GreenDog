#!/usr/bin/env python3
"""
Parse Med Contacts 1.xlsx and generate SQL INSERT statements for med_contacts table.
Uses the "Contacts" sheet (primary, more detailed) and "Updated Contacts Feb 2026" sheet (supplementary).
"""

import openpyxl
import json
import re

def clean(val):
    """Clean cell value"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', 'None', 'NA', 'N/A', 'NO PW LISTED', 'NO SITE LISTED'):
        return None
    return s

def escape_sql(val):
    """Escape single quotes for SQL"""
    if val is None:
        return 'NULL'
    return "'" + val.replace("'", "''") + "'"

def is_category_header(row_vals):
    """Check if a row is a category header (all caps vendor name, mostly empty other fields)"""
    vendor = clean(row_vals[0])
    if not vendor:
        return False
    # Category headers are typically ALL CAPS and have few other fields filled
    non_empty = sum(1 for v in row_vals[1:] if clean(v))
    return vendor == vendor.upper() and non_empty <= 1 and len(vendor) > 3

def parse_contacts_sheet(ws):
    """Parse the main Contacts sheet"""
    contacts = []
    current_category = 'General'
    current_vendor = None
    
    headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Headers: Vendor, Account Forms filled out, Misc Info, Contact, Email, Phone,
    #          Account #, Website, User ID, Password, Order Method, Payment Method,
    #          Notes, Issues, Statements, Itemized order emails, Orders
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        vendor = clean(vals[0]) if len(vals) > 0 else None
        misc_info = clean(vals[2]) if len(vals) > 2 else None
        contact = clean(vals[3]) if len(vals) > 3 else None
        email = clean(vals[4]) if len(vals) > 4 else None
        phone = clean(vals[5]) if len(vals) > 5 else None
        account = clean(vals[6]) if len(vals) > 6 else None
        website = clean(vals[7]) if len(vals) > 7 else None
        user_id = clean(vals[8]) if len(vals) > 8 else None
        password = clean(vals[9]) if len(vals) > 9 else None
        order_method = clean(vals[10]) if len(vals) > 10 else None
        payment_method = clean(vals[11]) if len(vals) > 11 else None
        notes = clean(vals[12]) if len(vals) > 12 else None
        issues = clean(vals[13]) if len(vals) > 13 else None
        
        # Combine notes and issues
        all_notes_parts = [n for n in [notes, issues] if n]
        all_notes = '; '.join(all_notes_parts) if all_notes_parts else None
        
        # Skip entirely empty rows
        row_vals = [vendor, contact, email, phone, account, website, user_id, password]
        if not any(row_vals):
            continue
        
        # Check if this is a category header
        if vendor and vendor == vendor.upper() and not any([contact, email, phone, website, user_id, password]):
            current_category = vendor
            continue
        
        # If vendor is present, it's a new vendor or sub-entry
        if vendor:
            current_vendor = vendor
            sub_label = None
        else:
            # Sub-entry under current vendor
            sub_label = contact if contact and not email and not phone else None
            if sub_label:
                # This is a role label like "Sherman Oaks Rep"
                pass
        
        # Determine the display vendor name
        display_vendor = current_vendor or 'Unknown'
        
        # Skip rows that only have empty category filler
        if not any([contact, email, phone, website, user_id, password, account]):
            if vendor and vendor != vendor.upper():
                current_vendor = vendor
            continue
        
        record = {
            'vendor_name': display_vendor,
            'category': current_category,
            'sub_label': misc_info or sub_label,
            'contact_name': contact,
            'contact_email': email,
            'contact_phone': phone,
            'account_number': str(account).rstrip('.0') if account else None,
            'website': website,
            'login_user_id': user_id,
            'login_password': password,
            'order_method': order_method,
            'payment_method': payment_method,
            'notes': all_notes,
        }
        
        contacts.append(record)
    
    return contacts

def parse_updated_sheet(ws):
    """Parse the Updated Contacts Feb 2026 sheet"""
    contacts = []
    
    # Headers: VENDOR, INFO, FORM FILLED, DATE FILLED, ACCT NO, LOCATION, DEPT, COSTS, CCFEES,
    #          WEBSITE, USER, PW, CONTACT NAME, CONTACT EM, CONTACT PH, CHROME/SAFARI, MISC, MISC
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        vendor = clean(vals[0]) if len(vals) > 0 else None
        info = clean(vals[1]) if len(vals) > 1 else None
        acct_no = clean(vals[4]) if len(vals) > 4 else None
        location = clean(vals[5]) if len(vals) > 5 else None
        dept = clean(vals[6]) if len(vals) > 6 else None
        website = clean(vals[9]) if len(vals) > 9 else None
        user_id = clean(vals[10]) if len(vals) > 10 else None
        password = clean(vals[11]) if len(vals) > 11 else None
        contact_name = clean(vals[12]) if len(vals) > 12 else None
        contact_email = clean(vals[13]) if len(vals) > 13 else None
        contact_phone = clean(vals[14]) if len(vals) > 14 else None
        browser = clean(vals[15]) if len(vals) > 15 else None
        misc1 = clean(vals[16]) if len(vals) > 16 else None
        misc2 = clean(vals[17]) if len(vals) > 17 else None
        
        if not vendor:
            continue
        
        # Skip header row and empty category rows
        if vendor == 'VENDOR' or (not any([website, user_id, password, contact_name, contact_email, contact_phone, acct_no]) and not info):
            continue
        
        # Skip OLD section
        if 'OLD' in (vendor or '') and 'NO LONGER' in (vendor or ''):
            break
        
        notes_parts = [n for n in [info, misc1, misc2] if n]
        notes = '; '.join(notes_parts) if notes_parts else None
        
        record = {
            'vendor_name': vendor,
            'category': 'Updated Feb 2026',
            'sub_label': info,
            'contact_name': contact_name,
            'contact_email': contact_email,
            'contact_phone': contact_phone,
            'account_number': str(acct_no).rstrip('.0') if acct_no else None,
            'website': website,
            'login_user_id': user_id,
            'login_password': password,
            'order_method': None,
            'payment_method': None,
            'notes': notes,
            'location': location,
            'department': dept,
            'browser_preference': browser,
        }
        
        contacts.append(record)
    
    return contacts

def deduplicate(contacts_main, contacts_updated):
    """Merge contacts, preferring main sheet data but adding unique entries from updated sheet"""
    # Use main sheet as primary
    all_contacts = list(contacts_main)
    
    # Track existing vendor+user combos from main sheet
    existing = set()
    for c in contacts_main:
        key = (c['vendor_name'], c.get('login_user_id', ''), c.get('contact_name', ''))
        existing.add(key)
    
    # Add unique entries from updated sheet
    for c in contacts_updated:
        key = (c['vendor_name'], c.get('login_user_id', ''), c.get('contact_name', ''))
        if key not in existing:
            all_contacts.append(c)
            existing.add(key)
    
    return all_contacts

def generate_sql(contacts):
    """Generate SQL INSERT statements"""
    lines = []
    lines.append("-- Auto-generated from Med Contacts 1.xlsx and Med Contacts 2.xlsx")
    lines.append("-- Generated on 2026-03-03")
    lines.append("")
    lines.append("INSERT INTO public.med_contacts (vendor_name, category, sub_label, contact_name, contact_email, contact_phone, account_number, website, login_user_id, login_password, order_method, payment_method, notes, location, department, browser_preference, is_active)")
    lines.append("VALUES")
    
    value_rows = []
    for c in contacts:
        row = "({}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, true)".format(
            escape_sql(c.get('vendor_name')),
            escape_sql(c.get('category')),
            escape_sql(c.get('sub_label')),
            escape_sql(c.get('contact_name')),
            escape_sql(c.get('contact_email')),
            escape_sql(c.get('contact_phone')),
            escape_sql(c.get('account_number')),
            escape_sql(c.get('website')),
            escape_sql(c.get('login_user_id')),
            escape_sql(c.get('login_password')),
            escape_sql(c.get('order_method')),
            escape_sql(c.get('payment_method')),
            escape_sql(c.get('notes')),
            escape_sql(c.get('location')),
            escape_sql(c.get('department')),
            escape_sql(c.get('browser_preference')),
        )
        value_rows.append(row)
    
    lines.append(',\n'.join(value_rows))
    lines.append(";")
    
    return '\n'.join(lines)

def main():
    wb = openpyxl.load_workbook('data/marketing/Med COntacts 1.xlsx', data_only=True)
    
    # Parse main Contacts sheet
    contacts_main = parse_contacts_sheet(wb['Contacts'])
    print(f"Parsed {len(contacts_main)} contacts from main Contacts sheet")
    
    # Parse Updated sheet
    contacts_updated = parse_updated_sheet(wb['Updated Contacts Feb 2026'])
    print(f"Parsed {len(contacts_updated)} contacts from Updated Contacts Feb 2026 sheet")
    
    # Deduplicate
    all_contacts = deduplicate(contacts_main, contacts_updated)
    print(f"Total unique contacts after dedup: {len(all_contacts)}")
    
    # Generate SQL
    sql = generate_sql(all_contacts)
    
    output_file = 'supabase/migrations/20260303000003_seed_med_contacts.sql'
    with open(output_file, 'w') as f:
        f.write(sql)
    
    print(f"SQL written to {output_file}")
    
    # Also output JSON for reference
    json_file = 'data/marketing/parsed/med_contacts_parsed.json'
    with open(json_file, 'w') as f:
        json.dump(all_contacts, f, indent=2, default=str)
    
    print(f"JSON written to {json_file}")
    
    # Summary by category
    categories = {}
    for c in all_contacts:
        cat = c.get('category', 'Unknown')
        categories[cat] = categories.get(cat, 0) + 1
    print("\nBy category:")
    for cat, count in sorted(categories.items()):
        print(f"  {cat}: {count}")

if __name__ == '__main__':
    main()
